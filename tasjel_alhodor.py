import os
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

# =========================================
# تحميل ملف .env المشترك
# =========================================
load_dotenv("/Users/samahsarsour/Desktop/mybot/.env")

BOT_TOKEN = os.getenv("ATTENDANCE_BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID", "0"))

# =========================================
# الملفات والأوراق
# =========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "attendance.xlsx")
COURSE_IMAGE = os.path.join(BASE_DIR, "course.png")

SHEET_ATTENDANCE = "attendance"
SHEET_STUDENTS_NEW = "students new"
SHEET_STUDENTS_OLD = "students old"

# =========================================
# حالة الجلسة الحالية
# =========================================
attendance = {
    "active": False,
    "records": [],
    "user_ids": set(),
    "started_at": None,
    "session_date": None,
    "message_chat_id": None,
    "message_id": None,
}


# =========================================
# دوال الوقت والمساعدة
# =========================================
def now_dt() -> datetime:
    return datetime.now()


def today_str() -> str:
    return now_dt().strftime("%d-%m-%Y")


def time_str() -> str:
    return now_dt().strftime("%I:%M %p")


def is_admin(user_id: int) -> bool:
    return user_id == ADMIN_ID


def normalize_name(name: str) -> str:
    return " ".join((name or "").strip().split()).casefold()


# =========================================
# إدارة ملف Excel
# =========================================
def ensure_workbook():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()

        ws_attendance = wb.active
        ws_attendance.title = SHEET_ATTENDANCE
        ws_attendance.append(["الاسم الثلاثي", "التاريخ", "وقت التسجيل", "الحالة"])

        ws_new = wb.create_sheet(SHEET_STUDENTS_NEW)
        ws_new.append(["الاسم الثلاثي"])

        ws_old = wb.create_sheet(SHEET_STUDENTS_OLD)
        ws_old.append(["الاسم الثلاثي"])

        wb.save(EXCEL_FILE)
        wb.close()
        return

    wb = load_workbook(EXCEL_FILE)

    if SHEET_ATTENDANCE not in wb.sheetnames:
        ws_attendance = wb.create_sheet(SHEET_ATTENDANCE)
        ws_attendance.append(["الاسم الثلاثي", "التاريخ", "وقت التسجيل", "الحالة"])

    if SHEET_STUDENTS_NEW not in wb.sheetnames:
        ws_new = wb.create_sheet(SHEET_STUDENTS_NEW)
        ws_new.append(["الاسم الثلاثي"])

    if SHEET_STUDENTS_OLD not in wb.sheetnames:
        ws_old = wb.create_sheet(SHEET_STUDENTS_OLD)
        ws_old.append(["الاسم الثلاثي"])

    wb.save(EXCEL_FILE)
    wb.close()


def read_names_from_sheet(sheet_name: str) -> list[str]:
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]

    names = []
    seen = set()

    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        full_name = row[0]
        if not full_name:
            continue

        full_name = " ".join(str(full_name).split()).strip()
        key = normalize_name(full_name)

        if key not in seen:
            names.append(full_name)
            seen.add(key)

    wb.close()
    return names


def get_all_students() -> list[str]:
    old_names = read_names_from_sheet(SHEET_STUDENTS_OLD)
    new_names = read_names_from_sheet(SHEET_STUDENTS_NEW)

    all_names = []
    seen = set()

    for name in old_names + new_names:
        key = normalize_name(name)
        if key not in seen:
            all_names.append(name)
            seen.add(key)

    return all_names


def student_exists(full_name: str) -> bool:
    target = normalize_name(full_name)
    return any(normalize_name(name) == target for name in get_all_students())


def add_student_to_new_sheet(full_name: str):
    if student_exists(full_name):
        return

    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_STUDENTS_NEW]
    ws.append([full_name])
    wb.save(EXCEL_FILE)
    wb.close()


def save_session_to_excel():
    ensure_workbook()

    session_date = attendance["session_date"]
    present_records = attendance["records"]
    all_students = get_all_students()

    present_map = {
        normalize_name(record["full_name"]): record
        for record in present_records
    }

    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_ATTENDANCE]

    written_present = set()
    for record in present_records:
        key = normalize_name(record["full_name"])
        if key in written_present:
            continue

        ws.append([
            record["full_name"],
            session_date,
            record["registered_time"],
            "حاضر",
        ])
        written_present.add(key)

    for student_name in all_students:
        key = normalize_name(student_name)
        if key not in present_map:
            ws.append([
                student_name,
                session_date,
                "",
                "لم يحضر",
            ])

    wb.save(EXCEL_FILE)
    wb.close()


# =========================================
# واجهة الرسائل
# =========================================
def build_button(active: bool = True):
    if active:
        return InlineKeyboardMarkup([
            [InlineKeyboardButton("اضغط هنا للتسجيل", callback_data="register")]
        ])

    return InlineKeyboardMarkup([
        [InlineKeyboardButton("تم إغلاق التسجيل", callback_data="closed")]
    ])


def build_text() -> str:
    date_text = attendance["session_date"] or today_str()
    start_time = attendance["started_at"].strftime("%I:%M %p") if attendance["started_at"] else "-"

    if not attendance["records"]:
        names_block = "لا يوجد تسجيل حتى الآن"
    else:
        lines = []
        for i, record in enumerate(attendance["records"], start=1):
            lines.append(f"{i}) {record['full_name']} | {record['registered_time']}")
        names_block = "\n".join(lines)

    return (
        "📋 كشف الحضور\n"
        "━━━━━━━━━━━━\n\n"
        f"📅 التاريخ: {date_text}\n"
        f"🕒 وقت البداية: {start_time}\n"
        f"👥 العدد: {len(attendance['records'])}\n\n"
        "━━━━━━━━━━━━\n"
        f"{names_block}"
    )


# =========================================
# أوامر البوت
# =========================================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "أهلاً بك 👋\n\n"
        "الأوامر المتاحة:\n"
        "/start_attendance - بدء تسجيل الحضور\n"
        "/show_attendance - عرض كشف الحضور الحالي\n"
        "/end_attendance - إنهاء التسجيل\n"
        "/myid - معرفة رقمك"
    )


async def myid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"🆔 رقمك هو:\n{update.effective_user.id}")


async def echo_private_for_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type == "private" and is_admin(update.effective_user.id):
        text = update.message.text.strip()
        await update.message.reply_text(f"وصلتني رسالتك ✅\n\n{text}")


async def start_attendance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("ليس لديك صلاحية.")
        return

    ensure_workbook()

    attendance["active"] = True
    attendance["records"] = []
    attendance["user_ids"] = set()
    attendance["started_at"] = now_dt()
    attendance["session_date"] = today_str()

    if os.path.exists(COURSE_IMAGE):
        with open(COURSE_IMAGE, "rb") as photo:
            sent = await update.message.reply_photo(
                photo=photo,
                caption=build_text(),
                reply_markup=build_button(True)
            )
    else:
        sent = await update.message.reply_text(
            build_text(),
            reply_markup=build_button(True)
        )

    attendance["message_chat_id"] = sent.chat_id
    attendance["message_id"] = sent.message_id


async def register(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user = query.from_user

    await query.answer()

    if query.data == "closed":
        await query.answer("تم إغلاق التسجيل.", show_alert=True)
        return

    if not attendance["active"]:
        await query.answer("التسجيل مغلق", show_alert=True)
        return

    if user.id in attendance["user_ids"]:
        await query.answer("تم تسجيلك مسبقًا", show_alert=True)
        return

    full_name = user.full_name or user.first_name or "بدون اسم"
    full_name = " ".join(full_name.split()).strip()

    add_student_to_new_sheet(full_name)

    attendance["user_ids"].add(user.id)
    attendance["records"].append({
        "full_name": full_name,
        "registered_time": time_str(),
    })

    try:
        if query.message.photo:
            await query.message.edit_caption(
                caption=build_text(),
                reply_markup=build_button(True)
            )
        else:
            await query.message.edit_text(
                text=build_text(),
                reply_markup=build_button(True)
            )
    except Exception:
        pass

    await query.answer("تم تسجيل حضورك ✅", show_alert=True)


async def show_attendance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(build_text())


async def end_attendance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("ليس لديك صلاحية.")
        return

    if not attendance["active"]:
        await update.message.reply_text("لا توجد جلسة حضور مفتوحة الآن.")
        return

    attendance["active"] = False

    save_session_to_excel()

    try:
        if attendance["message_chat_id"] and attendance["message_id"]:
            await context.bot.edit_message_reply_markup(
                chat_id=attendance["message_chat_id"],
                message_id=attendance["message_id"],
                reply_markup=build_button(False)
            )
    except Exception:
        pass

    await update.message.reply_text("تم إغلاق التسجيل ✅")

    if attendance["records"]:
        summary = "\n".join(
            f"{i+1}. {record['full_name']} - {record['registered_time']}"
            for i, record in enumerate(attendance["records"])
        )
    else:
        summary = "لا يوجد حضور مسجل."

    await context.bot.send_message(
        chat_id=ADMIN_ID,
        text=(
            "📋 الملخص النهائي\n\n"
            f"📅 التاريخ: {attendance['session_date']}\n"
            f"👥 العدد: {len(attendance['records'])}\n\n"
            f"{summary}"
        )
    )

    if os.path.exists(EXCEL_FILE):
        with open(EXCEL_FILE, "rb") as f:
            await context.bot.send_document(
                chat_id=ADMIN_ID,
                document=f,
                filename="attendance.xlsx",
                caption="📎 ملف الحضور المحدث"
            )


# =========================================
# التشغيل
# =========================================
def main():
    if not BOT_TOKEN:
        raise ValueError("ATTENDANCE_BOT_TOKEN غير موجود في ملف .env")

    ensure_workbook()

    app = ApplicationBuilder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("myid", myid))
    app.add_handler(CommandHandler("start_attendance", start_attendance))
    app.add_handler(CommandHandler("show_attendance", show_attendance))
    app.add_handler(CommandHandler("end_attendance", end_attendance))
    app.add_handler(CallbackQueryHandler(register, pattern="^(register|closed)$"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, echo_private_for_admin))

    print("Bot is running...")
    print("ADMIN_ID =", ADMIN_ID)
    print("EXCEL_FILE =", EXCEL_FILE)

    app.run_polling()


if __name__ == "__main__":
    main()
